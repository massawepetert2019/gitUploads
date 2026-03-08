from django.shortcuts import render
from django.http import JsonResponse,HttpResponse
from .models import MembersInfo
import openpyxl

# Create your views here.
def get_member_data(request):
	membId=request.GET.get('member_id')
	# print(membId)
	member=MembersInfo.objects.get(memberID=membId)
	# return HttpResponse(memberID)
	data={
		'cheo':member.Role,
		'phone':member.phone
	}
	return JsonResponse(data)

def AddMemberByForm(request):
	if request.method=="POST":
		fullname=request.POST.get('fullname')
		Uid=request.POST.get('memberID')
		phone=request.POST.get('phone')
		gender=request.POST.get('gender')
		cheo=request.POST.get('cheo')
		# if exixt		
		if MembersInfo.objects.filter(memberID=Uid).exists():			
			return JsonResponse({'error':'Samahani M-Koba ID Inatumika Na mwanachama Mwingine'})
		else:
			MembersInfo.objects.create(
				memberID=Uid,
				fullname=fullname,
				Role=cheo,
				phone=phone,
				gender=gender
				)		
			return JsonResponse({'success':'Umefanikiwa Kumsajili Mwanachama'})


def AddMemberByUpload(request):
	if request.method=="POST":
		# print('ddd')
		# return JsonResponse({'success':'ddddddddddddd'})
		try:
			excel_file = request.FILES.get('file_data')				
			if not excel_file:
				return JsonResponse({'success':'No file Selected'})

			wb=openpyxl.load_workbook(excel_file)
			sheet=wb.active
			for row in sheet.iter_rows(min_row=3,values_only=True):
				uid=row[0]
				fullname=row[1]
				gender=row[2]
				role=row[3]
				phone=row[4]
				# d=f'{uid}-{fullname}-{role}-{phone}'

				#avoid dublicate entry
				if not MembersInfo.objects.filter(memberID=uid).exists():
					
					MembersInfo.objects.create(
						memberID=uid,
						fullname=fullname,
						Role=role,
						phone=phone,
						gender=gender				
						)
			return JsonResponse({'success':'Uploaded Successfull' })

		except Exception as e:
			return JsonResponse({'success':'kuna tatizoo'})
		
		
			

		
		
		# for row in sheet.iter_rows(min_row=3,value_only=True):
		# 	uid=row[0]
		# 	fname=row[1]
		# 	# print(uid)
		# return JsonResponse({'success':row})

			#avoid dublicate entry
		# 	if not MembersInfo.objects.filter(memberID=Uid).exists():
		# 		MembersInfo.objects.create(
		# 		memberID=Uid,
		# 		fullname=fullname,
		# 		Role=cheo,
		# 		phone=phone,
		# 		gender=gender
		# 		)
		# return JsonResponse({'success':uid })




			
	

